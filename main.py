#!/usr/bin/python
# File: Epson_WS_V1.py
# Name: Samuel Guyette
# Desc: Pulls data from numerous websites to compare pricing of T-series competitors.
#		Will output a .csv file to location of program.
# Other files required: ws_functions.py, product_class.py, lists.py

import string
import datetime
import statistics

#xlsx
import os
import glob
import csv
import openpyxl
import pandas as pd

#import object class
from product_class import Product

#import helper function
from ws_functions import build_data_set

#import lists
from lists import *


def append_category(f, product_set, category, green_words, red_words):
	#picks correct category
	for product in product_set:
		write = True

		if product.added:
			write = False

		for word in red_words:
			if word in product.name:
				write = False

		if write:
			for substring in green_words:
				if substring in product.name and not product.added:
					f.write(product.channel + "," + product.country + "," + product.website + "," + product.company + "," + category + ",")
					f.write(product.name + "," + product.id + ", " + product.price + "," + product.shipping + "\n")
					product.added = True


def build_condensed_table(f, product_set):
	header = ",Website,"
	for i in sku_targets:
		header = header + i + ","
	f.write(header+"\n")
	for website in website_targets:
		#write country
		line = ""
		if "(CA)" in website:
			line += "CA,"
			website = website.replace(' (CA)','')
		else:
			if "\n" in website:
				pass
			else:
				line += "US,"
		line += website+","
		for sku in sku_targets:
			product_found = False
			for product in product_set:
				if product.id == sku and product.website == website and not product_found:
					line += product.price

					#add product to avg_price_hash
					num = True
					try:
						price = product.price
						try:
							price = price.split('.', 1)[0]
						except:
							pass
						try:
							price = price.replace('*','')
						except:
							pass

						price = price.replace('$','')
						price = int(price)
					except:
						num = False

					try:
						if num:
							avg_price_hash[sku].append(price)
					except:
						pass

					#check if lowest sales price is correct
					try:
						lsp = price_target_hash[sku]
						up = price_up_hash[sku]
						if product.country == "US" and product.website != "Epson":
							if price < lsp:
								line = line+" ↓"
							if price > up:
								line = line+" ↑"
					except:
						pass

					if "Free" in product.shipping:
						line += "*,"
					else:
						line += ","

					product_found = True

			if not product_found:
				line += ","

		f.write(line+"\n")


def find_averages(f, product_set):
	median_list = "\n,Median:,"
	mean_list = "\n,Mean:,"
	for sku in avg_price_hash:
		price_array = avg_price_hash[sku]
		median = statistics.median(price_array)
		mean = statistics.mean(price_array)
		median = round(median, 2)
		mean = round(mean, 2)
		median_list += str(median)+","
		mean_list += str(mean)+","

	f.write(median_list)
	f.write(mean_list+"\n")


def highlight_prices():
	#convert .csv file to .xlsv
	wb = openpyxl.Workbook()
	ws = wb.active
	with open("rough_output.csv", 'rt') as f:
		reader = csv.reader(f)
		for r, row in enumerate(reader, start=1):
			for c, val in enumerate(row, start=1):
				try:
					val = val.replace('$','')
					ws.cell(row=r, column=c).value = float(val)
				except:
					ws.cell(row=r, column=c).value = val

	wb.save("final_output.xlsx")

	#change letters to highlights
	df = pd.read_excel("final_output.xlsx")
	#chagne header location
	df.columns = df.iloc[2]
	df.reindex(df.index.drop(2))
	#writes over converted file
	fname = 'final_output.xlsx'
	writer = pd.ExcelWriter(fname, engine='xlsxwriter')
	df.to_excel(writer, sheet_name='MAIN', index=False)
	# get xlsxwriter objects
	workbook  = writer.book
	worksheet = writer.sheets['MAIN']

	#highlights incorrect prices
	formatH = workbook.add_format({'bg_color':'#FFC7CE', 'font_color':'#000000'})
	formatL = workbook.add_format({'bg_color':'#C6EFCE','font_color': '#000000'})
	worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
											'criteria': 'containing',
											'value':'↓',
											'format': formatH})
	worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
											'criteria': 'containing',
											'value':'↑',
											'format': formatL})

	writer.save()




# ****MAIN**** #
def main():
	#opens connection
	filename = "rough_output.csv"
	f = open(filename, "w")

	#hash set for all product objects
	product_set = set()
	build_data_set(product_set)
	print(str(len(product_set))+" potential products gathered.\n")
	#Writes name
	f.write("\nCreated by Sam Guyette\n")

	#Writes date and time of when ran
	now = datetime.datetime.now()
	f.write("Date and time of script execution: "+now.strftime("%Y-%m-%d %H:%M:%S")+"\n\n")

	#builds condensed table
	print("Building condensed table...\n")
	title1 = "Condensed Table\n*Free Shipping\n                            =  Price bellow lowest sales price"
	title2 = " permitted (LSPP): ↓\n                            =  Price above unilateral price: ↑\n\n"
	up_prices = ",UP:,"
	lsp = ",LSPP:,"
	for key in price_up_hash:
		up_prices = up_prices + str(price_up_hash[key]) + ","
	for key in price_target_hash:
		lsp = lsp + str(price_target_hash[key]) + ","

	f.write(title1)
	f.write(title2)
	f.write(up_prices+"\n")
	f.write(lsp+"\n\n")
	build_condensed_table(f, product_set)
	find_averages(f, product_set)
	space = "\n\n\n\n\n"
	f.write(space)

	#write headers for super table
	title = "Super Table\n"
	f.write(title)
	headers = "Channel, Country, Website, Company, Category, Name, SKU, Price, Shipping\n"
	f.write(headers)

	#creates sub categories
	print("Writing all data to super table...\n")
	#builds super table
	append_category(f, product_set, "Printer", printer_include_list, printer_exclude_list)
	append_category(f, product_set, "Ink", ink_include_list, ink_exclude_list)
	append_category(f, product_set, "Accessory", accessories_include_list, accessories_exlude_list)

	#highlights prices that are over or under recommended selling point
	print("Highlighting prices...\n")
	highlight_prices()

	#remove csv file
	print("Deleting csv file...\n")
	os.remove("rough_output.csv")

	print("Intern work is now complete.\n")

	f.close()


if __name__ == '__main__':
	main()













#
